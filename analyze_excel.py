import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('modelo_indicadores/g1_prazo_decorrido.xlsx')
    print(f"✅ Arquivo carregado com sucesso!")
    print(f"\n📋 Abas encontradas: {wb.sheetnames}")
    print(f"   Total de abas: {len(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"📄 ABA: {sheet_name}")
        print(f"{'='*60}")
        
        # Ler primeiras 15 linhas
        for i, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
            # Filtrar células vazias no final
            row_data = [cell for cell in row if cell is not None]
            if row_data:
                print(f"Linha {i}: {row_data}")
        
        print(f"\nTotal de linhas com dados: {sheet.max_row}")
        print(f"Total de colunas: {sheet.max_column}")
    
    wb.close()
    
except Exception as e:
    print(f"❌ Erro ao analisar arquivo: {e}")
    sys.exit(1)
